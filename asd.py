from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook


chrome = webdriver.Chrome()

chrome.get("https://www.saucedemo.com/")



name_user= chrome.find_element(By.ID, "user-name")
password=chrome.find_element(By.ID, "password")
buttom=chrome.find_element(By.XPATH, '//*[@id="login-button"]')


name_user.send_keys("standard_user")
password.send_keys("secret_sauce")
buttom.click()



all_name= chrome.find_elements(By.CLASS_NAME, "inventory_item_name")

all_description= chrome.find_elements(By.CLASS_NAME, "inventory_item_desc")

all_price= chrome.find_elements(By.CLASS_NAME, "inventory_item_price")



wb = Workbook()

ws = wb.active



ws['A1'] = "№"
ws['B1'] = "Name"
ws['C1'] = "Description"
ws['D1'] = "Price"

for i in range(len(all_name)):
    ws[f'A{i+2}'] = i+1
    ws[f'B{i+2}'] = all_name[i].text
    ws[f'C{i+2}'] = all_description[i].text
    ws[f'D{i+2}'] = all_price[i].text



wb.save("file2.xlsx")